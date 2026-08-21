import json
import re
import sys
import unicodedata
import warnings
import webbrowser
from pathlib import Path
from typing import Any

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import load_workbook

try:
    import config as cfg
except Exception:
    cfg = None

warnings.filterwarnings("ignore", message="Data Validation extension is not supported")


# =========================
# CONFIGURACION
# =========================

def get_base_path() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def cfg_value(name: str, default: Any) -> Any:
    return getattr(cfg, name, default) if cfg is not None else default


BASE_DIR = get_base_path()

EXCEL_FILE = Path(cfg_value("EXCEL_FILE", BASE_DIR / "Dashboard VT.xlsx"))
if not EXCEL_FILE.is_absolute():
    EXCEL_FILE = BASE_DIR / EXCEL_FILE

TEMPLATE_DIR = Path(cfg_value("TEMPLATE_DIR", cfg_value("TEMPLATE_PATH", BASE_DIR / "templates")))
if not TEMPLATE_DIR.is_absolute():
    TEMPLATE_DIR = BASE_DIR / TEMPLATE_DIR

OUTPUT_DIR = Path(cfg_value("OUTPUT_DIR", cfg_value("OUTPUT_PATH", BASE_DIR / "output")))
if not OUTPUT_DIR.is_absolute():
    OUTPUT_DIR = BASE_DIR / OUTPUT_DIR

TEMPLATE_NAME = cfg_value("TEMPLATE_VT", cfg_value("TEMPLATE_FILE", "dashboard_vt.html"))
OUTPUT_NAME = cfg_value("OUTPUT_VT", cfg_value("OUTPUT_FILE", "Dashboard_vt.html"))

SHEET_TABLAS = cfg_value("SHEET_TABLAS_GRAFICOS", "Tablas Gráficos")
SHEET_GASTOS_REALES = cfg_value("SHEET_GASTOS_REALES", "Gastos reales (SAP)")

MESES_DEFAULT = cfg_value(
    "MESES_DASHBOARD",
    [
        "ene-26", "feb-26", "mar-26", "abr-26", "may-26", "jun-26",
        "jul-26", "ago-26", "sep-26", "oct-26", "nov-26", "dic-26",
    ],
)

OPEN_BROWSER = bool(cfg_value("OPEN_BROWSER", True))


# =========================
# UTILIDADES
# =========================

def sin_acentos(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value: Any) -> str:
    text = sin_acentos(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def clean_label(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%b-%y").lower()
    return str(value).strip()


def to_number(value: Any) -> float:
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()

    if "#" in text:
        return 0.0

    text = text.replace("$", "").replace(" ", "")

    # Formato argentino: 1.234,56
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return 0.0


def fmt_fecha(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def fmt_dia(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d")
    return str(value).strip()


def used_rows(ws):
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def cell(row: list[Any], idx: int | None) -> Any:
    if idx is None:
        return None
    return row[idx] if 0 <= idx < len(row) else None


def find_sheet_name(wb, wanted: str) -> str:
    wanted_norm = norm(wanted)

    for name in wb.sheetnames:
        if norm(name) == wanted_norm:
            return name

    raise ValueError(
        f"No existe la hoja '{wanted}'. Hojas disponibles: {', '.join(wb.sheetnames)}"
    )


def find_row_index(rows: list[list[Any]], text: str) -> int | None:
    target = norm(text)

    for i, row in enumerate(rows):
        if any(norm(v) == target for v in row if v is not None):
            return i

    return None


def find_header_row(rows: list[list[Any]]) -> int:
    for i, row in enumerate(rows[:80]):
        values = [norm(v) for v in row]

        if "referencia" in values and "item" in values:
            return i

    raise ValueError("No se encontró el encabezado Referencia / Item en la hoja Tablas Gráficos.")


def month_columns(header_row: list[Any]) -> list[tuple[int, str]]:
    meses_norm = {norm(m): m for m in MESES_DEFAULT}
    cols = []

    for idx, value in enumerate(header_row):
        key = norm(value)

        if key in meses_norm:
            cols.append((idx, meses_norm[key]))

    if not cols:
        raise ValueError("No se encontraron columnas de meses en Tablas Gráficos.")

    return cols


# =========================
# LECTURA DE TABLAS GRAFICOS
# =========================

def build_base_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """
    Lee la tabla principal de Tablas Gráficos:
    Referencia | Item | ene-26 | feb-26 | ... | dic-26 | FY

    Devuelve filas normalizadas para el HTML:
    {referencia: 'BUD'|'Real', item: 'VT', mes: 'ene-26', valor: 123}
    """
    header_idx = find_header_row(rows)
    header = rows[header_idx]
    values = [norm(v) for v in header]

    ref_col = values.index("referencia")
    item_col = values.index("item")
    month_cols = month_columns(header)

    out = []

    for row in rows[header_idx + 1:]:
        ref = clean_label(cell(row, ref_col))
        item = clean_label(cell(row, item_col))
        ref_norm = norm(ref)

        # La base termina cuando dejan de aparecer filas BUD / Real.
        if not ref and not item:
            if out:
                break
            continue

        if ref_norm not in {"bud", "budget", "real"}:
            if out:
                break
            continue

        ref_out = "BUD" if ref_norm in {"bud", "budget"} else "Real"

        for col_idx, mes in month_cols:
            out.append({
                "referencia": ref_out,
                "item": item,
                "mes": mes,
                "valor": to_number(cell(row, col_idx)),
            })

    return out


def periodo_actual_from_base(base_rows: list[dict[str, Any]], meses: list[str]) -> str:
    for mes in reversed(meses):
        has_real = any(
            row["mes"] == mes
            and norm(row["referencia"]) == "real"
            and to_number(row["valor"]) != 0
            for row in base_rows
        )

        if has_real:
            return mes

    return meses[0] if meses else ""


# =========================
# LECTURA DE GASTOS REALES SAP
# =========================

def read_gastos_reales(wb) -> list[dict[str, Any]]:
    """
    Lee la hoja Gastos reales (SAP) y normaliza los datos para el dashboard.

    Columnas esperadas:
    - Período
    - Centro de coste
    - Clase de coste
    - Denom.clase de coste
    - Cantidad total reg.
    - Ud. cantidad contab.
    - Texto de pedido
    - Texto breve de material
    - Material
    - Valor/mon.inf.
    - Fecha de documento
    - Fe.contabilización
    - Usuario
    - Denominación
    - Referencia
    - Hora de entrada
    - Documento compras
    - Grupo Cta
    - Grupo Gral
    - UET
    - Taller
    """
    sheet_name = find_sheet_name(wb, SHEET_GASTOS_REALES)
    ws = wb[sheet_name]

    rows = list(used_rows(ws))
    if not rows:
        return []

    header = [norm(v) for v in rows[0]]

    def col(*names: str) -> int | None:
        names_norm = {norm(n) for n in names}

        for i, h in enumerate(header):
            if h in names_norm:
                return i

        return None

    c_periodo = col("Período", "Periodo")
    c_ceco = col("Centro de coste")
    c_clase = col("Clase de coste")
    c_denom_clase = col("Denom.clase de coste")
    c_cantidad = col("Cantidad total reg.")
    c_unidad = col("Ud. cantidad contab.")
    c_texto_pedido = col("Texto de pedido")
    c_texto_material = col("Texto breve de material")
    c_material = col("Material")
    c_importe = col("Valor/mon.inf.")
    c_fecha_doc = col("Fecha de documento")
    c_fecha_cont = col("Fe.contabilización", "Fe.contabilizacion")
    c_usuario = col("Usuario")
    c_denominacion = col("Denominación", "Denominacion")
    c_referencia = col("Referencia")
    c_doc_compras = col("Documento compras")
    c_grupo_cta = col("Grupo Cta")
    c_grupo_gral = col("Grupo Gral")
    c_uet = col("UET")
    c_taller = col("Taller")

    required = {
        "Período": c_periodo,
        "Centro de coste": c_ceco,
        "Valor/mon.inf.": c_importe,
        "Fe.contabilización": c_fecha_cont,
        "UET": c_uet,
        "Taller": c_taller,
    }

    missing = [name for name, idx in required.items() if idx is None]
    if missing:
        raise ValueError(
            f"Faltan columnas obligatorias en '{sheet_name}': {', '.join(missing)}"
        )

    out = []

    for row in rows[1:]:
        importe = to_number(cell(row, c_importe))

        # Si no hay importe, no aporta al ranking.
        if importe == 0:
            continue

        fecha_cont = cell(row, c_fecha_cont)
        periodo = cell(row, c_periodo)

        descripcion = (
            clean_label(cell(row, c_denominacion))
            or clean_label(cell(row, c_texto_pedido))
            or clean_label(cell(row, c_texto_material))
            or clean_label(cell(row, c_denom_clase))
        )

        out.append({
            "periodo": clean_label(periodo),
            "mes": clean_label(periodo),
            "dia": fmt_dia(fecha_cont),
            "fecha_documento": fmt_fecha(cell(row, c_fecha_doc)),
            "fecha_contabilizacion": fmt_fecha(fecha_cont),
            "ceco": clean_label(cell(row, c_ceco)),
            "clase_coste": clean_label(cell(row, c_clase)),
            "denom_clase_coste": clean_label(cell(row, c_denom_clase)),
            "cantidad": to_number(cell(row, c_cantidad)),
            "unidad": clean_label(cell(row, c_unidad)),
            "texto_pedido": clean_label(cell(row, c_texto_pedido)),
            "texto_material": clean_label(cell(row, c_texto_material)),
            "material": clean_label(cell(row, c_material)),
            "descripcion": descripcion,
            "importe": importe,
            "importe_abs": abs(importe),
            "usuario": clean_label(cell(row, c_usuario)),
            "referencia": clean_label(cell(row, c_referencia)),
            "documento_compras": clean_label(cell(row, c_doc_compras)),
            "grupo_cta": clean_label(cell(row, c_grupo_cta)),
            "grupo_gral": clean_label(cell(row, c_grupo_gral)),
            "uet": clean_label(cell(row, c_uet)),
            "taller": clean_label(cell(row, c_taller)),
        })

    return out


# =========================
# CONSTRUCCION DE DATOS
# =========================

def construir_datos_graficos() -> dict[str, Any]:
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"No se encontró el Excel: {EXCEL_FILE}")

    wb = load_workbook(EXCEL_FILE, data_only=True, read_only=True)

    sheet_tablas_name = find_sheet_name(wb, SHEET_TABLAS)

    ws = wb[sheet_tablas_name]
    rows = list(used_rows(ws))

    base = build_base_rows(rows)
    meses = [m for m in MESES_DEFAULT if any(r["mes"] == m for r in base)]

    if not meses:
        meses = list(MESES_DEFAULT)

    periodo_actual = periodo_actual_from_base(base, meses)
    gastos = read_gastos_reales(wb)

    datos = {
        "meses": meses,
        "periodo_actual": periodo_actual,
        "base": base,
        "gastos": gastos,
    }

    print("Datos cargados:")
    print(f"- Filas base: {len(datos['base'])}")
    print(f"- Meses: {', '.join(datos['meses'])}")
    print(f"- Período actual: {datos['periodo_actual']}")
    print(f"- Gastos reales SAP: {len(datos['gastos'])} filas")

    return datos


# =========================
# RENDER HTML
# =========================

def find_template_dir() -> Path:
    candidates = [
        TEMPLATE_DIR / TEMPLATE_NAME,
        BASE_DIR / TEMPLATE_NAME,
    ]

    for path in candidates:
        if path.exists():
            return path.parent

    searched = "\n".join(f"- {p}" for p in candidates)
    raise FileNotFoundError(f"No se encontró el template {TEMPLATE_NAME}. Buscado en:\n{searched}")


def render_dashboard(datos_graficos: dict[str, Any]) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    template_dir = find_template_dir()

    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(["html", "xml"]),
    )

    template = env.get_template(TEMPLATE_NAME)
    html = template.render(datos_graficos=datos_graficos)

    output_path = OUTPUT_DIR / OUTPUT_NAME
    output_path.write_text(html, encoding="utf-8")

    # Archivo de control para revisar qué datos llegaron al HTML.
    debug_path = OUTPUT_DIR / "datos_graficos_debug.json"
    debug_path.write_text(json.dumps(datos_graficos, ensure_ascii=False, indent=2), encoding="utf-8")

    return output_path


def abrir_html(path: Path) -> None:
    if OPEN_BROWSER and path.exists():
        webbrowser.open(path.resolve().as_uri())


def run_actualizacion() -> Path:
    datos_graficos = construir_datos_graficos()
    html_path = render_dashboard(datos_graficos)

    print(f"Dashboard VT generado: {html_path}")

    return html_path


if __name__ == "__main__":
    try:
        html_generado = run_actualizacion()
        abrir_html(html_generado)
    except Exception as exc:
        print(f"Error generando Dashboard VT: {exc}")
        raise